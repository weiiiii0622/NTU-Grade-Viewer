import asyncio
from decimal import ROUND_HALF_UP, Decimal, getcontext
from functools import reduce
import math
import os
import pickle
import re
from enum import Enum
from pathlib import Path
from typing import Union

from aiohttp import ClientSession
from models import GRADES, Course, GradeElement, GradeWithUpdate, Segment, UpdateBase
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from tqdm.asyncio import tqdm_asyncio
from routes.submit import insert_grades
from utils.general import extract_dict
from utils.grade import get_segments
from utils.search import Key, search_course
from utils.segment_list import SegmentList

# ----------------------------- Parse excel data ----------------------------- #

MAX = len(GRADES) - 1
MIN = 0

TOTAL = Decimal(100)

WIDTH = 8
GRADES_HEADER = "".join(f"{x.ljust(WIDTH)}" for x in reversed(GRADES))


def extend_segments(segments: list[Segment]) -> list[Segment]:
    seglist: SegmentList = SegmentList(MAX + 1, TOTAL, None, strict=False)
    for seg in segments:
        seglist.remove(*seg.unpack())

    dump = seglist.dump()
    if len(dump) == 1:
        # if there is only one segment left, `value` can be inferred
        l, r, _ = dump[0]
        new_seg = Segment(l=l, r=r, value=TOTAL - sum(seg.value for seg in segments))
        left = (seg for seg in segments if seg.l < new_seg.l)
        right = (seg for seg in segments if seg.l > new_seg.l)

        return [*left, new_seg, *right]

    elif len(dump) > 1:
        # just drop this grade
        return []
    else:
        return segments


# @dataclass
# class ParsedGradeInfo:
#     title: str
#     lecturer: str
#     semester: Annotated[str, Field(pattern=r"\d+-\d+")]
#     segments: list[Segment]
#     cls: str = field(default="")

#     def __post_init__(self):


#         # ? maybe we leave this part to server side
#         validate_segments(self.segments)

# def readable_str(self):
#     total = 19
#     info = 10
#     title = self.title[: total - info]
#     title = title.ljust(total - info, "　")
#     s = title + f"　{self.lecturer}　{self.semester}"[:info].ljust(info, "　") + " | "
#     seg_idx = len(self.segments) - 1
#     for i in range(MAX, MIN - 1, -1):
#         seg = self.segments[seg_idx]
#         if i == seg.r:
#             val = f"{seg.value:.2%}"
#             if seg.l == seg.r:
#                 s += val.ljust(WIDTH)
#             else:
#                 s += f"({val.center(WIDTH*len(seg)-3)}) "
#             seg_idx -= 1
#     # s += "\n"
#     return s


def validate_segments(segments: list[Segment]):
    for i in range(len(segments) - 1):
        assert segments[i].r + 1 == segments[i + 1].l, f"{segments[i].r}, {segments[i+1].l}"
    assert math.isclose(sum(seg.value for seg in segments), TOTAL, abs_tol=1), sum(
        seg.value for seg in segments
    )


def extract_cls(raw_lecturer: str):
    """
    Used to seperate str like "林明仁(02)" to "林明仁", "02"
    """
    if obj := re.match(r"(.+).*?\((.+?)\)", raw_lecturer):
        return obj.group(1).strip(), obj.group(2)
    return raw_lecturer, ""


# * Color
class Color(Enum):
    BLACK = 0
    BLUE = 1
    ORANGE = 2


def getColor(cell: Cell):
    theme_map = {1: 0, 4: 1, 8: 2}
    if cell.font.color.theme:
        try:
            return Color(theme_map[int(cell.font.color.theme)])
        except:
            pass

    color_map = {
        "FF000000": 0,
        "FF4285F4": 1,
        "FFFF6D01": 2,
        "FFFF9900": 2,
        "FFED7D31": 2,
    }
    if cell.font.color.rgb:
        return Color(color_map[cell.font.color.rgb])

    else:
        print(cell)
        print(cell.font.color)
        raise Exception()


stop = False


# * Row parser
async def parse_row(row: tuple[Cell, ...], session: ClientSession) -> Union[GradeElement, None]:

    title, lecturer, semester = [str(c.value).strip() for c in row[:3]]
    title = title.replace("(", "（")
    title = title.replace(")", "）")
    lecturer, class_id = extract_cls(lecturer.replace("（", "(").replace("）", ")"))
    lecturer = lecturer.replace(" ", "")

    segments: list[Segment] = []

    st = -1
    end = -1
    blue_val: Decimal = Decimal(-1)

    # * Note the row is reversed
    for i, c in enumerate(reversed(row[3:13])):
        if c.value:
            val = Decimal(str(c.value))
            assert (
                type(c.value) == float or type(c.value) == int
            ), f"{c}, {c.value}: {type(c.value)}"

            match getColor(c):
                case Color.BLACK:
                    if st != -1:
                        segments.append(Segment(l=st, r=end, value=blue_val))
                        st = -1
                        end = -1
                        blue_val = Decimal(-1)
                    segments.append(Segment(l=i, r=i, value=val))
                case Color.BLUE:
                    if st == -1:
                        st = i
                    end = i
                    print(c.value, val)
                    blue_val = val
                case Color.ORANGE:
                    assert end == -1
                    if segments:
                        # this should be "higher than"
                        segments.append(Segment(l=i, r=MAX, value=val))
                    else:
                        # "lower than"
                        segments.append(Segment(l=MIN, r=i, value=val))
    assert st == -1 and end == -1, f"{st}, {end}, {segments}"

    total = sum(seg.value for seg in segments)
    if math.isclose(total, 1, abs_tol=Decimal("0.02")) or any(
        seg.value <= Decimal("0.1") for seg in segments
    ):
        for seg in segments:
            seg.value *= 100

    segments = extend_segments(segments)
    if not segments:
        return None

    res = await search_course(
        extract_dict(["title", "lecturer", "semester", "class_id"], locals()),  # type: ignore
        thresholds={"class_id": 5, "lecturer": 1},  # class_id may be missing in data
        session=session,
    )
    global stop
    if stop:
        return

    keys = ["title", "lecturer", "class_id", "semester"]
    if not res:
        return
    # assert res

    # if not class_id and res["class_id"]:
    #     class_id = res["class_id"]

    keys: list[Key] = ["title", "lecturer", "class_id", "semester"]
    cur = extract_dict(keys, locals())

    # * Manually check if any case other than typo
    if any(res[k] != cur[k] for k in keys):
        diffs: list[Key] = [k for k in keys if res[k] != cur[k]]
        for k in diffs:
            if k == "title" and cur["title"].replace("（", "(").replace("）", ")") == res["title"]:
                continue
            if k == "title" and cur["title"].replace("：", ":") == res["title"]:
                continue
            # print(k)
            # print(f"{res[k]} != {cur[k]}")

    # ? special case, just give up
    if title == "健康體適能" and not class_id:
        return

    segments = extend_segments(segments)
    # ? this will be done by `GradeElement`
    # validate_segments(segments)

    # trust the result from `search_course`
    course = Course(**extract_dict(["id1", "id2", "title"], res))
    grade = GradeElement(
        course=course,
        course_id1=course.id1,
        course_id2=course.id2,
        segments=segments,
        **extract_dict(["lecturer", "class_id", "semester"], res),
    )
    return grade


# ----------------------------------- Test ----------------------------------- #


async def get_grades() -> list[GradeElement]:

    data_dir = Path(__file__).parent / "../../data/pre-collected/"

    wb = load_workbook(str(data_dir / "raw/106-110學年NTU課程成績比例.xlsx"))
    ws = wb.worksheets[0]
    rows = list(ws.rows)

    rows = [row for row in rows if row[0].value]

    async with ClientSession() as session:
        grades = await tqdm_asyncio.gather(*(parse_row(row, session) for row in rows[1:]))
        # grades = await asyncio.gather(*(parse_row(row, session) for row in rows[1:]))

    valid_grades: list[GradeElement] = [g for g in grades if g]
    for g in valid_grades:
        if len([_g for _g in valid_grades if _g.id == g.id]) > 1:
            assert False, "duplicate course"

    print(f"drop {len(grades)-len(valid_grades)} / {len(grades)}")
    # TODO: insert into db
    # TODO: add `confirmed` field in db
    return valid_grades


# todo: fix weird data


async def main():
    if os.path.exists("grades.tmp"):
        with open("grades.tmp", "rb") as f:
            grades = pickle.load(f)
    else:
        grades = await get_grades()
        with open("grades.tmp", "wb+") as f:
            pickle.dump(grades, f)
    print(f"{len(grades)} grades!")

    def convert_to_udpates(g: GradeElement) -> list[GradeWithUpdate]:
        results: list[GradeWithUpdate] = []
        prev_sum = Decimal(0)
        try:
            for seg in g.segments:
                if seg.l == seg.r:
                    update = UpdateBase(
                        pos=seg.l, lower=prev_sum, higher=TOTAL - seg.value - prev_sum, solid=False
                    )
                    results.append(GradeWithUpdate(**g.model_dump(), update=update))
                prev_sum += seg.value
        except Exception as e:
            print(e)
            print(g)
            print(results)
            print(prev_sum)
        # try:
        #     assert g.segments == get_segments([_g.update for _g in results])
        # except:
        #     print(g)
        #     print(results)
        #     print()
        #     exit(0)
        return results

    # grade_updates = [convert_to_udpates(g) for g in grades]
    grades_updates = reduce(
        lambda prev, next: prev + next, [convert_to_udpates(g) for g in grades], []
    )

    await insert_grades(grades=grades_updates)


asyncio.run(main())

data_dir = Path(__file__).parent / "../../data/pre-collected/"

wb = load_workbook(str(data_dir / "raw/106-110學年NTU課程成績比例.xlsx"))
ws = wb.worksheets[0]
rows = list(ws.rows)


f = lambda x: x.quantize(Decimal(".00"), ROUND_HALF_UP)
# class A:
#     a: Decimal
